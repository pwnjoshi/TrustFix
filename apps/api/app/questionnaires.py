from __future__ import annotations

import csv
import io
from typing import Any
from openpyxl import load_workbook, Workbook

from .models import ReviewQuestion


CANDIDATE_NAMES = {"question", "security question", "requirement", "questionnaire item", "control question"}


class QuestionnaireError(ValueError):
    pass


def _detect(headers: list[str], requested: str | None = None) -> int:
    normalized = [str(h).strip().lower() for h in headers]
    if requested:
        try:
            return normalized.index(requested.strip().lower())
        except ValueError as exc:
            raise QuestionnaireError(f"Question column '{requested}' was not found") from exc
    for idx, value in enumerate(normalized):
        if value in CANDIDATE_NAMES or "question" in value:
            return idx
    raise QuestionnaireError("TrustFix could not identify a question column. Select the column explicitly.")


def parse_csv(content: bytes, question_column: str | None = None) -> tuple[list[ReviewQuestion], list[str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise QuestionnaireError("CSV must use UTF-8 encoding") from exc
    rows = list(csv.reader(io.StringIO(text)))
    if len(rows) < 2:
        raise QuestionnaireError("Questionnaire must contain a header and at least one question")
    headers = rows[0]
    index = _detect(headers, question_column)
    questions = [ReviewQuestion(original_row=row_number, question=row[index].strip()) for row_number, row in enumerate(rows[1:], 2) if len(row) > index and row[index].strip()]
    if not questions:
        raise QuestionnaireError("No non-empty questions were found")
    return questions, headers


def parse_xlsx(content: bytes, question_column: str | None = None) -> tuple[list[ReviewQuestion], list[str]]:
    try:
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = book.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception as exc:
        raise QuestionnaireError("The XLSX workbook could not be read") from exc
    if len(rows) < 2:
        raise QuestionnaireError("Questionnaire must contain a header and at least one question")
    headers = [str(value or "") for value in rows[0]]
    index = _detect(headers, question_column)
    questions = [ReviewQuestion(original_row=row_number, question=str(row[index]).strip()) for row_number, row in enumerate(rows[1:], 2) if len(row) > index and row[index] is not None and str(row[index]).strip()]
    return questions, headers


def export_csv(review) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Question", "Answer", "Status", "Evidence summary", "Verified at", "Reviewer"])
    for item in sorted(review.questions, key=lambda q: q.original_row):
        writer.writerow([item.question, item.answer or "", item.status or "Unmapped", "; ".join(item.evidence_ids), "", ""])
    return buffer.getvalue().encode("utf-8-sig")


def export_xlsx(review) -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = "TrustFix responses"
    sheet.append(["Question", "Answer", "Status", "Evidence summary", "Verified at", "Reviewer"])
    for item in sorted(review.questions, key=lambda q: q.original_row):
        sheet.append([item.question, item.answer or "", str(item.status or "Unmapped"), "; ".join(item.evidence_ids), "", ""])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 72
    sheet.column_dimensions["B"].width = 60
    output = io.BytesIO()
    book.save(output)
    return output.getvalue()

